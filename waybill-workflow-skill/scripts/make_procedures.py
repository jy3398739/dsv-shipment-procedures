#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
订舱邮件 -> 电池货出货手续包 生成器

流程：
  1. 解析订舱邮件(.eml)，提取 主单号 / 分单号(HAWB) / 件重 / 航班 / 含电池机型
  2. 用 机型 去 苹果手机-网站查询信息汇总.xlsx（chinasdg 官网查询结果）匹配对应鉴定，取「物品名称（中文）」(含电池参数)
  3. 对每个有模板(模板/<主单>*.*) 的主单：
       - 填 交运单.xlsx（表样）-> 新文件
       - 校验 托书/安检单/应急措施(PDF) + 托运人声明(docx) 是否含 主单号/机型/航班
       - 打包成 出货手续包/<主单>/

用法：
  python make_procedures.py <订舱邮件.eml> [--template-dir ../模板] [--xlsx 苹果手机-网站查询信息汇总.xlsx] [--out-dir .]
"""
import argparse, email, glob, html, json, os, re, shutil, sys, zipfile

MODEL_ORDER = ["A3090", "A2638", "A3102", "A3106", "A3296", "A3287", "A2890", "A3409"]

FW2HW = {ord(c): ord(h) for c, h in zip("　（）：，、；", " () :,;")}
def norm(s):
    if s is None:
        return ""
    s = re.sub(r"<[^>]+>", "", str(s))
    s = html.unescape(s).translate(FW2HW)
    return re.sub(r"\s+", "", s)


# ---------- 1. 解析订舱邮件 ----------
def parse_email(path):
    msg = email.message_from_binary_file(open(path, "rb"))
    body = None
    for part in msg.walk():
        if part.get_content_type() == "text/plain":
            body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", "ignore")
            break
    if not body:
        return []
    headers = [m.start() for m in re.finditer(r"件数\s*计费\s*毛重", body)]
    blocks = []
    positions = sorted(headers)
    for i, h in enumerate(positions):
        end = positions[i + 1] if i + 1 < len(positions) else len(body)
        blocks.append(body[h:end])

    masters = []
    for b in blocks:
        m = re.search(r"\d{3}-\d{8}", b)
        if not m:
            continue
        master = m.group(0)
        flight = None
        fm = re.search(r"EK\d{3}", b)
        if fm:
            flight = fm.group(0)
        hawb_lines = []
        battery_models = set()
        for line in b.splitlines():
            hm = re.search(r"TYN\d+", line)
            if not hm:
                continue
            nums = re.findall(r"\d+(?:\.\d+)?", line)
            pcs = int(nums[0]) if len(nums) >= 1 else None
            wt = float(nums[1]) if len(nums) >= 2 else None
            is_battery = "手机含电池" in line
            models = re.findall(r"A\d{4}", line)
            hawb_lines.append({"hawb": hm.group(0), "pcs": pcs, "wt": wt,
                               "models": models, "battery": is_battery})
            if is_battery:
                battery_models.update(models)
        masters.append({
            "master": master, "flight": flight, "hawbs": hawb_lines,
            "models": sorted(battery_models, key=lambda x: MODEL_ORDER.index(x) if x in MODEL_ORDER else 99),
        })
    return masters


# ---------- 2. 加载 chinasdg 鉴定 ----------
def load_chinasdg(xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["官网查询结果"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    ci_code = next(i for i, h in enumerate(hdr) if h and "报告编号" in str(h))
    ci_cn = next(i for i, h in enumerate(hdr) if h and "中文" in str(h))
    ci_en = next(i for i, h in enumerate(hdr) if h and "英文" in str(h))
    model2cert = {}
    for r in rows[1:]:
        cn = norm(r[ci_cn])
        mm = re.search(r"A\d{4}", cn)
        if not mm:
            continue
        model = mm.group(0)
        bat = re.search(r"(?:APN:?\s*)?([A-Za-z0-9]+),\s*([\d.]+)V", r[ci_cn]) or \
              re.search(r"([A-Za-z0-9]+),\s*([\d.]+)V,\s*([\d.]+Wh)", r[ci_cn])
        model2cert[model] = {
            "code": r[ci_code],
            "cn": html.unescape(re.sub(r"<[^>]+>", "", str(r[ci_cn]))).strip(),
            "en": html.unescape(re.sub(r"<[^>]+>", "", str(r[ci_en]))).strip(),
        }
    return model2cert


# ---------- 3a. 填 交运单 ----------
def _find_row_by_col_value(ws, col, predicate):
    """返回列 col 中第一个满足 predicate(值) 的行号，找不到返回 None。"""
    for row in ws.iter_rows():
        for c in row:
            if c.column == col and c.value is not None and predicate(str(c.value).strip()):
                return c.row
    return None


def fill_jiaoyun(master, tpl_pdf_dir, data, out_dir):
    import openpyxl
    xlsx_list = glob.glob(os.path.join(tpl_pdf_dir, "*交运单.xlsx"))
    if not xlsx_list:
        return None, "无交运单模板"
    src = xlsx_list[0]
    wb = openpyxl.load_workbook(src)
    ws = wb["表样"]
    total_pcs = sum(h["pcs"] or 0 for h in data["hawbs"])
    total_wt = sum(h["wt"] or 0 for h in data["hawbs"])
    goods = [model2cert_global[m]["cn"] for m in data["models"] if m in model2cert_global]
    goods_text = "\n".join(goods)
    # 主单数据行：列B(2) 为 主单号格式，绝不碰表头
    master_row = _find_row_by_col_value(ws, 2, lambda v: re.fullmatch(r"\d{3}-\d{8}", v))
    if master_row:
        ws.cell(row=master_row, column=3, value="BOM")                       # 目的站
        ws.cell(row=master_row, column=4, value=total_pcs)                  # 件数
        ws.cell(row=master_row, column=5, value=round(total_wt, 1))          # 重量
        fc = ws.cell(row=master_row, column=6, value=goods_text)             # 品名
        fc.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    # 分单行：列B(2) 为 HAWB，精确匹配后填 目的站/件数/重量
    for h in data["hawbs"]:
        hr = _find_row_by_col_value(ws, 2, lambda v: v == h["hawb"])
        if hr:
            ws.cell(row=hr, column=3, value="BOM")
            ws.cell(row=hr, column=4, value=h["pcs"])
            ws.cell(row=hr, column=5, value=round(h["wt"] or 0, 1))
    out = os.path.join(out_dir, f"交运单_{data['master']}.xlsx")
    wb.save(out)
    return out, f"主单件数{total_pcs}/重量{round(total_wt,1)}kg，品名{len(goods)}条"


# ---------- 3b. 校验 PDF/docx 手续 ----------
def validate_pdf(path, master, models, flight):
    import fitz
    doc = fitz.open(path)
    txt = doc[0].get_text()
    doc.close()
    comp = norm(txt)
    miss_models = [m for m in models if m not in comp]
    return {
        "master": master in txt,
        "models_ok": not miss_models,
        "miss_models": miss_models,
        "flight": (flight in txt) if flight else None,
    }


def validate_docx(path, master, apns):
    z = zipfile.ZipFile(path)
    xml = z.read("word/document.xml").decode("utf-8", "ignore")
    txt = re.sub(r"<[^>]+>", "", xml)
    comp = norm(txt)
    miss = [a for a in apns if a not in comp]
    return {"master": master in txt, "apns_ok": not miss, "miss_apns": miss}


def apns_of(models):
    out = []
    for m in models:
        c = model2cert_global.get(m)
        if c:
            mm = re.search(r"(?:APN:?\s*)?([A-Za-z0-9]+),\s*([\d.]+)V", c["cn"]) or \
                 re.search(r"([A-Za-z0-9]+),\s*([\d.]+)V,\s*([\d.]+Wh)", c["cn"])
            if mm:
                out.append(mm.group(1))
    return out


def main():
    global model2cert_global
    ap = argparse.ArgumentParser()
    ap.add_argument("email")
    ap.add_argument("--template-dir", default="../模板")
    ap.add_argument("--xlsx", default=os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "..", "templates",
        "苹果手机-网站查询信息汇总.xlsx"))
    ap.add_argument("--out-dir", default=".")
    a = ap.parse_args()

    tpl_dir = os.path.abspath(a.template_dir)
    out_dir = os.path.abspath(a.out_dir)
    os.makedirs(out_dir, exist_ok=True)

    print("解析订舱邮件:", a.email)
    masters = parse_email(a.email)
    if not masters:
        print("!! 未从邮件解析到主单，退出")
        sys.exit(1)
    for m in masters:
        print(f"  主单 {m['master']} | 航班 {m['flight']} | 含电池机型 {m['models']}")

    print("\n加载 chinasdg 鉴定:", a.xlsx)
    model2cert_global = load_chinasdg(a.xlsx)
    print(f"  鉴定机型映射: {list(model2cert_global.keys())}")

    generated = []
    for m in masters:
        tpls = glob.glob(os.path.join(tpl_dir, f"{m['master']}*"))
        if not tpls:
            print(f"\n[跳过] 主单 {m['master']} 无模板")
            continue
        pkg = os.path.join(out_dir, "出货手续包", m["master"])
        os.makedirs(pkg, exist_ok=True)
        print(f"\n========== 主单 {m['master']} ==========")

        # 交运单
        jy, info = fill_jiaoyun(m, tpl_dir, m, out_dir)
        if jy:
            shutil.copy2(jy, os.path.join(pkg, os.path.basename(jy)))
            print(f"  [生成] 交运单: {jy} ({info})")
            generated.append(("交运单", jy))
        else:
            print(f"  [跳过] 交运单: {info}")

        # 校验 + 打包 PDF/docx
        apns = apns_of(m["models"])
        for p in tpls:
            base = os.path.basename(p)
            if base.lower().endswith(".pdf"):
                if "运单复本" in base:
                    continue  # 运单副本由 booking_to_waybill 单独生成
                r = validate_pdf(p, m["master"], m["models"], m["flight"])
                status = "OK" if (r["master"] and r["models_ok"] and (r["flight"] is not False)) else "FAIL"
                print(f"  [校验] {base}: 主单{r['master']} 机型{'OK' if r['models_ok'] else '缺'+str(r['miss_models'])} 航班{r['flight']} -> {status}")
                shutil.copy2(p, os.path.join(pkg, base))
                generated.append((base, os.path.join(pkg, base)))
            elif base.lower().endswith(".docx"):
                r = validate_docx(p, m["master"], apns)
                status = "OK" if (r["master"] and r["apns_ok"]) else "FAIL"
                print(f"  [校验] {base}: 主单{r['master']} 电池型号{'OK' if r['apns_ok'] else '缺'+str(r['miss_apns'])} -> {status}")
                shutil.copy2(p, os.path.join(pkg, base))
                generated.append((base, os.path.join(pkg, base)))

    print("\n========== 生成汇总 ==========")
    for name, path in generated:
        print(f"  {name}: {path}")


if __name__ == "__main__":
    main()
