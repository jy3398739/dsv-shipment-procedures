# -*- coding: utf-8 -*-
"""订舱邮件 -> 运单副本 生成流水线。

流程：
  1. 解析 .eml 订舱邮件，按 BOM 区块提取每个主单号、分单号(HAWB)、含电池 HAWB 的机型(电池型号)
  2. 读取 苹果手机-网站查询信息汇总.xlsx（chinasdg 官网查询结果）表，建立 机型 -> 鉴定(报告编号/物品名称中英文) 映射
  3. 对每个主单的机型匹配对应鉴定，标注未匹配机型
  4. 若主单存在模板文件「<主单> 运单复本.pdf」，用匹配到的鉴定品名生成运单副本；
     否则跳过并提示（等用户后续提供模板）

用法：
  python booking_to_waybill.py <邮件.eml> [--xlsx 苹果手机-网站查询信息汇总.xlsx] [--out-dir .]
"""
import argparse, email, glob, json, os, re, subprocess, sys, html, random

PY = sys.executable


def strip_html(s):
    if s is None:
        return ""
    return html.unescape(re.sub(r"<[^>]+>", "", str(s))).strip()


def read_body(path):
    msg = email.message_from_binary_file(open(path, "rb"))
    for part in msg.walk():
        if part.get_content_type() == "text/plain":
            return part.get_payload(decode=True).decode(
                part.get_content_charset() or "utf-8", "ignore")
    return None


def top_masters(body):
    """主单范围 = 第一个表格之前摘要里列出的主单（转发「完整版」里的其他票不处理）。"""
    h = re.search(r"件数\s+计费\s+毛重", body)
    head = body[: h.start()] if h else body
    return [re.sub(r"\s", "", m) for m in re.findall(r"\d{3}-\s*\d{8}", head)]


def parse_email(path):
    """返回 [ {master, hawb, pcs, wt, is_battery, models, desc} ]。"""
    body = read_body(path)
    if not body:
        return []

    # 按表头行「件数 计费 毛重」切分成 BOM 区块
    headers = [m.start() for m in re.finditer(r"件数\s+计费\s+毛重", body)]
    blocks = []
    for i, h in enumerate(headers):
        end = headers[i + 1] if i + 1 < len(headers) else len(body)
        blocks.append(body[h:end])

    recs = []
    master_re = re.compile(r"\d{3}-\s*\d{8}")   # 横线后可能有空格：176- 61333915
    hawb_re = re.compile(r"(TYN\d+)")
    # 件数/计费/毛重 三元组，后接尺寸括号行作锚（避免误配预报单号/航班行；尺寸行行首允许空格）
    triple_re = re.compile(r"(\d+)\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)\s*\n\s*\(")
    for blk in blocks:
        ms = master_re.findall(blk)
        master = re.sub(r"\s", "", ms[-1]) if ms else None  # 段内最后一个主单号
        prev = 0
        hms = list(hawb_re.finditer(blk))
        for i, hm in enumerate(hms):
            prefix = blk[prev:hm.start()]
            prev = hm.end()
            hawb = hm.group(1)
            tr = triple_re.findall(prefix)
            pcs = int(tr[-1][0]) if tr else None
            wt = float(tr[-1][2]) if tr else None
            # 区域止于下一个 HAWB（报关行不一定是泽坤，不能按泽坤切）
            nxt = hms[i + 1].start() if i + 1 < len(hms) else len(blk)
            region = blk[hm.end():nxt]
            is_battery = bool(re.search(r"手机(?:模组)?含电池", region))
            cat = ("module" if re.search(r"手机模组含电池", region)
                   else "phone" if is_battery else None)
            # 机型在「含电池」标记后的描述里，无分号也兼容：直接取全部 A\d{4}
            models = list(dict.fromkeys(re.findall(r"[A-Z]\d{4}", region))) if is_battery else []
            # 非电池分单=零件行：带 HS 编码的行取 (中文品名, 英文品名) 对
            # 中文=首个中文词组；英文=SVC … 标记，无则取行末纯字母段大写（speaker→SPEAKER）
            parts = []
            if not is_battery:
                seen_cn = set()
                for ln in region.split("\n"):
                    if "HS" not in ln:
                        continue
                    mcn = re.search(r"[一-鿿]{2,}", ln)
                    if not mcn or mcn.group(0) in seen_cn:
                        continue
                    cn = mcn.group(0)
                    men = re.search(r"SVC [A-Z0-9]+(?: [A-Z0-9]+)*", ln)
                    if men:
                        en = men.group(0)
                    else:
                        en = ""
                        for t in reversed([t.strip() for t in ln.replace("，", ",").split(",")]):
                            if t and re.fullmatch(r"[A-Za-z ]+", t):
                                en = t.upper()
                                break
                    seen_cn.add(cn)
                    parts.append([cn, en])
            recs.append({
                "master": master,
                "hawb": hawb,
                "pcs": pcs,
                "wt": wt,
                "is_battery": is_battery,
                "category": cat,
                "models": models,
                "parts": parts,
                "desc": re.sub(r"\s+", " ", region).strip()[:120],
            })
    return recs


def load_certs(xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    name = next((n for n in ("官网查询结果", "网站查询信息") if n in wb.sheetnames),
                wb.sheetnames[0])
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    def col(*keys):
        for k in keys:
            for i, h in enumerate(hdr):
                if h and k in str(h):
                    return i
        raise KeyError(keys)
    ci_code = col("报告编号", "证书编号")
    ci_cn = col("物品名称", "中文")
    ci_en = col("Name of Goods", "英文")
    certs = []
    for r in rows[1:]:
        if r[ci_cn]:
            cn = strip_html(r[ci_cn])
            en = strip_html(r[ci_en])
            models = re.findall(r"[A-Z]\d{4}", cn)
            certs.append({"code": r[ci_code], "cn": cn, "en": en, "models": models})
    return certs


def find_tpl(template_dir, master):
    hit = []
    for dp, _, fns in os.walk(template_dir):
        for f in fns:
            if f.startswith(master) and f.endswith(".pdf") and ("运单副本" in f or "运单复本" in f):
                hit.append(os.path.join(dp, f))
    return hit


def build_masters(recs, scope=None):
    """把 recs 聚成 gen_procedures 用的 MASTERS 结构：{master: {models, hawbs}}。
    转发邮件下方常带引用历史表格：无主单号的块丢弃，重复分单以首次出现为准；
    scope 给定时只聚合范围内的主单（= 邮件开头摘要列出的票）。"""
    masters = {}
    seen_hawb = set()
    for r in recs:
        if not r["master"]:
            continue
        if scope is not None and r["master"] not in scope:
            continue
        ent = masters.setdefault(r["master"],
                                 {"models": [], "hawbs": [], "parts": [], "bpcs": 0})
        if r["hawb"] in seen_hawb:
            continue
        if r["hawb"] and r["pcs"] is not None:
            seen_hawb.add(r["hawb"])
            ent["hawbs"].append((r["hawb"], r["pcs"], r["wt"]))
            if r["is_battery"]:
                ent["bpcs"] += r["pcs"]
        for mdl in r["models"]:
            if mdl not in ent["models"]:
                ent["models"].append(mdl)
        for cn, en in r.get("parts", []):
            if cn not in [x[0] for x in ent["parts"]]:
                ent["parts"].append([cn, en])
    return masters


def derive_battery(gp, models):
    """从 CN_GOODS 品名串反提内置电池 APN（托运人声明用）。"""
    bats = []
    for m in models:
        cn = gp.CN_GOODS.get(m)
        if not cn:
            print(f"  ⚠ 机型 {m} 不在 CN_GOODS，需先补鉴定品名再生成")
            continue
        bm = re.search(r"锂离子电池\s*(?:APN:?\s*)?([A-Z0-9]{2,10})", cn)
        if bm and bm.group(1) not in bats:
            bats.append(bm.group(1))
    return bats


def variant_cat(cn):
    """按去空白后的品名前缀判变体类别：phone=数字式手机/5G 变体，module=手机模组，None=其他。
    xlsx 里出现过「5G 数字移动电话机」（5G 与数字间带空格），严格 startswith 会漏判。"""
    s = re.sub(r"\s+", "", str(cn or ""))
    if s.startswith(("数字式手机", "5G数字")):
        return "phone"
    if s.startswith("手机模组"):
        return "module"
    return None


def cert_for(certs, mdl, cat):
    """按邮件分类挑 xlsx 行：phone→数字式手机/5G 变体；module→手机模组变体。
    页面手工补的证书（manual）不按前缀筛，两类变体都命中。"""
    for c in certs:
        if mdl not in c["models"]:
            continue
        if c.get("manual"):
            return c
        if cat == "phone" and variant_cat(c["cn"]) == "phone":
            return c
        if cat == "module" and variant_cat(c["cn"]) == "module":
            return c
    return None


def build_entries(recs, scope, certs, overrides=None):
    """每主单的品名条目 [(model, category, cert)]：按 HAWB 行的手机/模组分类各取对应变体，
    两属机型两类各一条；scope 外的主单不处理。
    overrides：{机型: 证书} 人工指定映射（核对单改过证书编号），优先于 xlsx 自动匹配。"""
    out, seen = {}, {}
    for r in recs:
        if not r["master"] or not r["is_battery"]:
            continue
        if scope is not None and r["master"] not in scope:
            continue
        lst = out.setdefault(r["master"], [])
        se = seen.setdefault(r["master"], set())
        for mdl in r["models"]:
            key = (mdl, r["category"])
            if key in se:
                continue
            se.add(key)
            c = (overrides or {}).get(mdl) or cert_for(certs, mdl, r["category"])
            lst.append((mdl, r["category"], c))
    return out


def derive_battery_from(cn_list):
    """从品名串列表反提内置电池 APN（托运人声明用），模组/手机变体都适用。"""
    bats = []
    for cn in cn_list:
        bm = re.search(r"锂离子电池\s*(?:APN:?\s*)?([A-Z0-9]{2,10})", cn)
        if bm and bm.group(1) not in bats:
            bats.append(bm.group(1))
    return bats


def inject_goods(gp, certs):
    """机型不在 CN_GOODS/EN_GOODS 时，按 xlsx 的「数字式手机/5G数字移动电话机」变体自动补。"""
    for c in certs:
        if variant_cat(c["cn"]) == "phone":
            for mdl in c["models"]:
                gp.CN_GOODS.setdefault(mdl, c["cn"])
                gp.EN_GOODS.setdefault(mdl, c["en"])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("eml")
    ap.add_argument("--xlsx", default=os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "..", "templates",
        "苹果手机-网站查询信息汇总.xlsx"),
        help="chinasdg 官网查询结果汇总（默认 templates/苹果手机-网站查询信息汇总.xlsx）")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--template-dir",
                    default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "templates"),
                    help="运单模板所在目录（递归搜索 运单副本/运单复本）")
    ap.add_argument("--random", action="store_true",
                    help="随机生成新运单号；默认保留原主单号（不覆盖模板，输出加『_副本』后缀）")
    ap.add_argument("--full-package", action="store_true",
                    help="邮件进→手续包出：提取结果聚 MASTERS，调 gen_procedures 每主单生成六份")
    a = ap.parse_args()

    body = read_body(a.eml)
    recs = parse_email(a.eml)
    certs = load_certs(a.xlsx)
    scope = set(top_masters(body)) if body else None

    model2cert = {}
    # 证书编号优先取「数字式手机/5G」变体行（与单据品名文字配对），无手机行才回退
    for c in certs:
        if variant_cat(c["cn"]) == "phone":
            for mdl in c["models"]:
                model2cert[mdl] = c
    for c in certs:
        for mdl in c["models"]:
            model2cert.setdefault(mdl, c)

    masters = build_masters(recs, scope)

    print("========== 1) 邮件订舱提取 ==========")
    for r in recs:
        tag = "含电池" if r["is_battery"] else "零件/模组"
        mdl = ",".join(r["models"]) if r["models"] else "-"
        print(f"  主单 {r['master']} | 分单 {r['hawb']} | {tag} | 件数 {r['pcs']} 毛重 {r['wt']} | 机型 {mdl}")

    print("\n========== 2) 主单 -> chinasdg 鉴定匹配 ==========")
    for master, ent in masters.items():
        matched = sorted(m for m in ent["models"] if m in model2cert)
        unmatched = sorted(m for m in ent["models"] if m not in model2cert)
        print(f"\n主单 {master}:")
        print(f"  含电池机型: {ent['models']}")
        if matched:
            print("  已匹配鉴定: " + ", ".join(f"{m}->{model2cert[m]['code']}" for m in matched))
        else:
            print("  (无)")
        if unmatched:
            print(f"  ⚠ 未匹配机型(无对应鉴定): {unmatched}")

    if a.full_package:
        print("\n========== 3) 邮件进 → 手续包出 (gen_procedures) ==========")
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        import gen_procedures as gp
        inject_goods(gp, certs)
        # 与内置 MASTERS 交叉核对，验证提取质量
        for master, ent in masters.items():
            bi = gp.MASTERS.get(master)
            if not bi:
                print(f"  ⚠ {master} 不在内置 MASTERS（新票，直接用提取结果）")
                continue
            same_m = sorted(ent["models"]) == sorted(bi["models"])
            same_h = ent["hawbs"] == bi["hawbs"]
            print(f"  {master}: 机型与内置一致={same_m} HAWB/件重一致={same_h}")
            if not same_m:
                print(f"    提取={sorted(ent['models'])}\n    内置={sorted(bi['models'])}")
            if not same_h:
                print(f"    提取={ent['hawbs']}\n    内置={bi['hawbs']}")
        entries_by = build_entries(recs, scope, certs)
        gp.MASTERS = {}
        for m, ent in masters.items():
            entries = [(c["cn"], c["en"]) for _, _, c in entries_by.get(m, []) if c]
            gp.MASTERS[m] = {
                "models": ent["models"],
                "entries": entries,
                "parts": ent.get("parts", []),
                "bpcs": ent.get("bpcs", 0),
                "battery": derive_battery_from([c for c, _ in entries]),
                "pcs": sum(h[1] for h in ent["hawbs"]),
                "hawbs": ent["hawbs"]}
        for master, info in gp.MASTERS.items():
            print(f"---- {master} ----")
            gp.make_shengming(master, info)
            gp.make_anxian(master, info)
            gp.make_yingji(master, info)
            gp.fill_tuoshū(master, info)
            gp.fill_waybill(master, info)
            gp.fill_jiaoyun(master, info)
            print(f"  六份完成 -> {gp.OUT}/{master}")
        print("ALL DONE (full-package)")
        return

    generated = []
    for master, ent in masters.items():
        matched = [m for m in ent["models"] if m in model2cert]
        tpl = find_tpl(a.template_dir, master)
        if not tpl:
            print(f"  ⚠ 主单 {master} 无模板（{master}*运单副本/复本.pdf），跳过生成（待用户提供）")
            continue

        # 去重组装品名
        seen, items = set(), []
        for m in matched:
            c = model2cert[m]
            if c["cn"] not in seen:
                seen.add(c["cn"])
                items.append((c["cn"], c["en"]))
        # 默认保留原主单号；--random 时随机生成
        new_no = (f"176-{random.randint(10000000, 99999999):08d}"
                  if a.random else master)
        ij = os.path.join(a.out_dir, "_items.json")
        with open(ij, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)
        # 新号==主单号时，输出加『_副本』后缀避免覆盖原模板
        if new_no == master:
            out = os.path.join(a.out_dir, f"{master} 运单复本_副本.pdf")
        else:
            out = os.path.join(a.out_dir, f"{new_no} 运单复本.pdf")
        mk = os.path.join(os.path.dirname(os.path.abspath(__file__)), "make_waybill.py")
        cmd = [PY, mk, "--template", tpl[0],
               "--old", master, "--new", new_no, "--items-json", ij,
               "--no-ai-label", "--out", out]
        print(f"  生成命令: {' '.join(cmd)}")
        subprocess.run(cmd, check=True)
        generated.append((master, new_no, out, len(items)))
        try:
            if os.path.exists(ij):
                os.remove(ij)
        except OSError:
            pass

    print("\n========== 4) 生成结果 ==========")
    for master, new_no, out, n in generated:
        note = "（保留原主单号）" if new_no == master else ""
        print(f"  主单 {master} -> 运单 {new_no}{note} ({n} 个品名): {out}")

    if not generated:
        print("  （无可用模板，未生成任何运单副本）")


if __name__ == "__main__":
    main()
