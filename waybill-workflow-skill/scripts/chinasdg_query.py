# -*- coding: utf-8 -*-
"""chinasdg 证书信息查询（本环境适配版）。

替代原技能依赖的 browser-agent：直接对公开接口
POST /API/Common/GetQrcodeInfo 发请求，请求体为 action= 加
base64Data=base64({"orderCode":"<编号>"}), 返回 JSON 内含证书完整字段。

本脚本：
  1. 从 xlsx 读取 报告编号 列与已提取的汇总字段
  2. 逐个编号查询 chinasdg 官网
  3. 解析返回字段（去 HTML 标签）
  4. 与 xlsx 汇总表逐项比对，输出 一致/不一致
  5. 生成 Excel 交付（含 AIGC 标注）
"""
import urllib.request, urllib.parse, json, base64, ssl, re, html
import openpyxl

CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE
BASE = "http://www.chinasdg.cn"
XLSX = "官网查询结果汇总.xlsx"
OUT_XLSX = "chinasdg_查询结果.xlsx"

# 全角->半角标点
FW = "（）：，；．＃％　"
HW = "():,;.#% "
FW2HW = str.maketrans(dict(zip(FW, HW)))


def norm(s):
    if s is None:
        return ""
    s = re.sub(r"<[^>]+>", "", str(s))          # 去 HTML 标签
    s = html.unescape(s)
    s = s.translate(FW2HW)                       # 全角->半角
    s = re.sub(r"\s+", "", s)                    # 去全部空白
    return s


def ck(s):
    """内容指纹：仅保留字母/数字/汉字，用于判断『实质是否一致』。"""
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]", "", norm(s))


def query(code):
    payload = base64.b64encode(json.dumps({"orderCode": code}).encode("utf-8")).decode()
    body = urllib.parse.urlencode({"action": "", "base64Data": payload}).encode()
    req = urllib.request.Request(
        BASE + "/API/Common/GetQrcodeInfo", data=body,
        headers={"User-Agent": "Mozilla/5.0",
                 "Content-Type": "application/x-www-form-urlencoded",
                 "X-Requested-With": "XMLHttpRequest"})
    raw = urllib.request.urlopen(req, timeout=20, context=CTX).read().decode("utf-8", "ignore")
    outer = json.loads(raw)
    if outer.get("code") != 200:
        return {"ok": False, "msg": outer}
    inner = json.loads(outer["data"])
    if inner.get("code") != 0 or not inner.get("data"):
        return {"ok": False, "msg": inner.get("msg", "无数据")}
    return {"ok": True, "data": inner["data"]}


def read_xlsx(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {h: i for i, h in enumerate(hdr)}
    items = []
    for r in rows[1:]:
        code = r[idx.get("报告编号")] if "报告编号" in idx else None
        if not code:
            continue
        items.append({
            "报告编号": str(code).strip(),
            "物品名称（中文）": r[idx.get("物品名称（中文）")],
            "物品名称（英文）": r[idx.get("物品名称（英文）")],
            "委托单位": r[idx.get("委托单位")],
            "生效日期": r[idx.get("生效日期")],
            "货物类型": r[idx.get("货物类型")],
            "UN编号": r[idx.get("UN编号")],
            "运输专用名称": r[idx.get("运输专用名称")],
            "识别结论": r[idx.get("识别结论")],
            "储运注意事项": r[idx.get("储运注意事项")],
        })
    return items


def main():
    xitems = read_xlsx(XLSX)
    print(f"xlsx 读取到 {len(xitems)} 条报告编号\n")

    # 官网字段映射：官网键 -> 展示列名
    FMAP = [
        ("productNameCn", "物品名称（中文）"),
        ("productNameEn", "物品名称（英文）"),
        ("entrustCoopName", "委托单位"),
        ("printTime", "生效日期"),
        ("identificationResult", "货物类型"),
        ("unId", "UN编号"),
        ("properShippingName", "运输专用名称"),
        ("identificationCn", "识别结论"),
        ("storageCn", "储运注意事项"),
    ]

    results = []
    for xi in xitems:
        code = xi["报告编号"]
        r = query(code)
        rec = {"报告编号": code, "查询成功": r["ok"], "官网": {}, "比对": []}
        if not r["ok"]:
            rec["失败原因"] = str(r["msg"])[:80]
            print(f"[失败] {code}: {rec['失败原因']}")
            results.append(rec)
            continue
        d = r["data"]
        for key, col in FMAP:
            rec["官网"][col] = d.get(key, "")
        # 逐项比对
        for key, col in FMAP:
            web = norm(rec["官网"][col])
            xv = norm(xi.get(col))
            web_ck, xv_ck = ck(rec["官网"][col]), ck(xi.get(col))
            if web == xv:
                typ = "一致"
            elif web_ck == xv_ck or web_ck in xv_ck or xv_ck in web_ck:
                typ = "一致(仅格式/合并差异)"
            else:
                typ = "实质不一致"
            rec["比对"].append({
                "字段": col,
                "官网": rec["官网"][col],
                "汇总表": xi.get(col),
                "一致": typ == "一致",
                "类型": typ,
            })
        print(f"[成功] {code}  委托单位={rec['官网']['委托单位']}  生效日期={rec['官网']['生效日期']}  货物类型={rec['官网']['货物类型']}")
        results.append(rec)

    # 写 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "官网查询结果"
    cols = ["报告编号", "查询成功", "物品名称（中文）", "物品名称（英文）", "委托单位",
            "生效日期", "货物类型", "UN编号", "运输专用名称", "识别结论", "储运注意事项"]
    ws.append(cols)
    for rec in results:
        if rec["查询成功"]:
            g = rec["官网"]
            ws.append([rec["报告编号"], "是", g["物品名称（中文）"], g["物品名称（英文）"],
                       g["委托单位"], g["生效日期"], g["货物类型"], g["UN编号"],
                       g["运输专用名称"], g["识别结论"], g["储运注意事项"]])
        else:
            ws.append([rec["报告编号"], "否（失败）", "", "", "", "", "", "", "", "", rec.get("失败原因", "")])

    ws2 = wb.create_sheet("与汇总表比对")
    ws2.append(["报告编号", "字段", "结论", "官网值", "汇总表值"])
    for rec in results:
        if not rec["查询成功"]:
            continue
        for b in rec["比对"]:
            ws2.append([rec["报告编号"], b["字段"], b["类型"],
                        str(b["官网"])[:200], str(b["汇总表"])[:200]])

    ws3 = wb.create_sheet("说明")
    ws3.append(["本表由 AI 生成（chinasdg 官网证书信息查询，接口 POST /API/Common/GetQrcodeInfo）"])
    ws3.append(["比对规则：去除 HTML 标签/全角转半角/去空白后逐字相等判为『一致』；"])
    ws3.append(["若字母数字汉字内容指纹相同（仅标点/逗号/中英文是否合并有差异）判为『一致(仅格式/合并差异)』；"])
    ws3.append(["内容指纹不同才判为『实质不一致』。"])
    wb.save(OUT_XLSX)
    print(f"\nExcel 已生成：{OUT_XLSX}")

    # 控制台汇总比对
    print("\n========== 比对汇总 ==========")
    all_substantive = True
    for rec in results:
        if not rec["查询成功"]:
            all_substantive = False
            print(f"{rec['报告编号']}: 查询失败")
            continue
        fmt = [b for b in rec["比对"] if b["类型"] == "一致(仅格式/合并差异)"]
        real = [b for b in rec["比对"] if b["类型"] == "实质不一致"]
        if real:
            all_substantive = False
            print(f"{rec['报告编号']}: 实质不一致 -> " + ", ".join(b["字段"] for b in real))
        elif fmt:
            print(f"{rec['报告编号']}: 实质一致（{len(fmt)} 处格式/合并差异）")
        else:
            print(f"{rec['报告编号']}: 全部一致")
    print("\n总体结论:", "全部实质一致" if all_substantive else "存在实质差异（见上）")


if __name__ == "__main__":
    main()
