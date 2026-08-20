# -*- coding: utf-8 -*-
"""一键出货手续包入口（脱离 AI 独立运行版）。

用法：
  把订舱邮件 .eml 拖到「一键出货手续包.bat」上，或双击 bat 用文件选择器选邮件。
  流程：解析校验 → （有缺失时开补充页补齐）→ 出「生成前核对单」等人工确认，
  核对单上可直接修改机型对应的鉴定证书编号 → 点「确认生成」才生成七份手续，
  生成后再出带逐项校验结果的正式核对单；取消则不生成，退出码非 0。
  设环境变量 RUNPKG_AUTO_CONFIRM=1 可跳过确认页直接生成（自动化/测试用）。

闸门（G1-G6）：
  G1 邮件能解析出表格块、主单号、HAWB
  G2 每个 HAWB 有件数/毛重；含电池 HAWB 有机型
  G3 目的站/航班/运输日期能从邮件提取
  G4 全部机型在鉴定汇总 xlsx 有对应证书
  G5 全部机型在 CN_GOODS/EN_GOODS 有品名变体；主单有 托书/运单副本 模板
  G6 生成后文本层校验（机型齐全、不超栏、占位符剥离、交运单数值）
"""
import os
import re
import sys
import email
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import booking_to_waybill as bw
import gen_procedures as gp

ROOT = os.path.join(HERE, "..", "..", "templates")
XLSX = os.path.join(ROOT, "苹果手机-网站查询信息汇总.xlsx")
OUT = os.path.normpath(os.path.join(HERE, "..", "..", "出货手续包"))

MONTHS = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
          'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}


def get_body(path):
    msg = email.message_from_binary_file(open(path, "rb"))
    for part in msg.walk():
        if part.get_content_type() == "text/plain":
            return part.get_payload(decode=True).decode(
                part.get_content_charset() or "utf-8", "ignore"), msg
    return None, msg


def extract_meta(body, msg):
    """目的站/航班/承运人/运输日期；返回 (meta, issues)。"""
    issues = []
    meta = {}
    m = re.search(r"(?m)^([A-Z]{3})\s*$", body)
    if m:
        meta["dest"] = m.group(1)
    else:
        issues.append("G3 目的站：邮件里没有独立成行的三字母目的站（如 RUH），请人工确认后在脚本里补。")
    m = re.search(r"([A-Z]{2}\d{3})\s*/\s*([A-Z][a-z]{2})\s+(\d{1,2})", body)
    if m:
        meta["flight"] = m.group(1)
        meta["carrier"] = m.group(1)[:2]
        mon = MONTHS.get(m.group(2))
        day = int(m.group(3))
        year = None
        if msg.get("Date"):
            ym = re.search(r"20\d{2}", msg["Date"])
            year = int(ym.group(0)) if ym else None
        if not year:
            ym = re.search(r"(20\d{2})", body)
            year = int(ym.group(1)) if ym else datetime.date.today().year
        if mon and day:
            meta["date"] = f"{year}-{mon:02d}-{day:02d}"
        else:
            issues.append("G3 运输日期：航班后的日期（如 Aug 16）解析失败，请人工填写。")
    else:
        issues.append("G3 航班：邮件里没有「EK309/Aug 16」样式的航班/日期，请人工确认。")
    return meta, issues


def verify_master(master, info):
    """生成后文本层校验，返回 [(检查项, 是否通过, 说明)]。"""
    import fitz
    import zipfile
    import openpyxl
    chk = []
    d = os.path.join(OUT, master)
    files = sorted(os.listdir(d)) if os.path.isdir(d) else []
    chk.append(("手续包文件齐全(7份)", len(files) == 7,
                "" if len(files) == 7 else ",".join(files)))

    f = next((x for x in files if "安检单" in x), None)
    if f:
        pg = fitz.open(os.path.join(d, f))[0]
        txt = pg.get_text().replace("\n", "")
        miss = [m for m in info["models"] if m not in txt]
        chk.append(("安检单机型齐全", not miss, ",".join(miss)))
        parts = info.get("parts") or []
        miss_p = [c for c, e in parts if c not in txt]
        chk.append(("安检单零件品名齐全", not miss_p, ",".join(miss_p)))
        ws_ = [b for b in pg.get_text("words") if 240 < b[1] < 410]
        mx = max((b[2] for b in ws_), default=0)
        my = max((b[3] for b in ws_), default=0)
        chk.append(("安检单品名不超右栏(x≤205.5)", mx <= 205.5, f"x1={mx:.1f}"))
        chk.append(("安检单品名不超下界(y≤405)", my <= 405, f"y1={my:.1f}"))
        chk.append(("安检单航班/日期/目的站已写",
                    gp.FLIGHT in txt and gp.DATE_STR in txt and gp.DEST in txt, ""))
    else:
        chk.append(("安检单存在", False, ""))

    f = next((x for x in files if "应急措施" in x), None)
    if f:
        txt = fitz.open(os.path.join(d, f))[0].get_text().replace("\n", "")
        miss = [m for m in info["models"] if m not in txt]
        chk.append(("应急措施机型齐全", not miss, ",".join(miss)))
    else:
        chk.append(("应急措施存在", False, ""))

    f = next((x for x in files if "托书" in x), None)
    if f:
        pg = fitz.open(os.path.join(d, f))[0]
        txt = pg.get_text()
        miss = [m for m in info["models"] if m not in txt]
        chk.append(("托书机型齐全", not miss, ",".join(miss)))
        if info.get("bpcs"):
            t2 = "".join(txt.split())
            chk.append((f"托书 LITHIUM BATTERY {info['bpcs']} PCS",
                        f"LITHIUMBATTERY{info['bpcs']}PCS" in t2, ""))
        ys = [s["origin"][1] for bl in pg.get_text("dict")["blocks"]
              for ln in bl.get("lines", []) for s in ln["spans"] if s["size"] < 6]
        ok = ys and min(ys) >= 360 and max(ys) <= 473
        chk.append(("托书品名不压标题/底线(360~473)", bool(ok),
                    "" if ok else f"y {min(ys):.1f}~{max(ys):.1f}"))
    else:
        chk.append(("托书存在", False, ""))

    f = next((x for x in files if "运单副本" in x), None)
    if f:
        txt = fitz.open(os.path.join(d, f))[0].get_text().replace("\n", "")
        miss = [m for m in info["models"] if txt.count(m) < 2]
        chk.append(("运单副本机型齐全(中英两段)", not miss, ",".join(miss)))
        chk.append(("运单副本 PI967/主单号", "PI 967" in txt and master in txt, ""))
        if info.get("bpcs"):
            t2 = txt.replace(" ", "")
            chk.append((f"运单副本 LITHIUM BATTERY {info['bpcs']} PCS",
                        f"LITHIUMBATTERY{info['bpcs']}PCS" in t2, ""))
    else:
        chk.append(("运单副本存在", False, ""))

    f = next((x for x in files if "托运人声明" in x), None)
    if f:
        xml = zipfile.ZipFile(os.path.join(d, f)).read("word/document.xml").decode("utf-8")
        chk.append(("托运人声明占位符已替换", "单击此处输入文字" not in xml, ""))
        chk.append(("托运人声明 sdt 已剥离", "<w:sdt>" not in xml, ""))
        chk.append(("托运人声明主单/件数", master in xml and str(info["pcs"]) in xml, ""))
    else:
        chk.append(("托运人声明存在", False, ""))

    f = next((x for x in files if "危险品确认单" in x), None)
    if f:
        xml = zipfile.ZipFile(os.path.join(d, f)).read("word/document.xml").decode("utf-8")
        goods = "，".join(gp.goods_of(info)[0])
        chk.append(("危险品确认单已填(主单/品名)",
                    master in xml and goods in xml, ""))
        miss = [c for c in (info.get("cert_codes") or []) if str(c) not in xml]
        chk.append(("危险品确认单报告编号齐全", not miss, ",".join(map(str, miss))))
    else:
        chk.append(("危险品确认单存在", False, ""))

    f = next((x for x in files if "交运单" in x), None)
    if f:
        ws = openpyxl.load_workbook(os.path.join(d, f))["表样"]
        wt = round(sum(h[2] for h in info["hawbs"]), 1)
        ok = (ws["B3"].value == master and ws["C3"].value == gp.DEST
              and ws["D3"].value == info["pcs"]
              and abs((ws["E3"].value or 0) - wt) < 0.01
              and ws["G3"].value == f"{len(info['hawbs'])}分")
        chk.append(("交运单主单行(目的站/件数/重量/N分)", bool(ok),
                    "" if ok else f"B3={ws['B3'].value} D3={ws['D3'].value} E3={ws['E3'].value} G3={ws['G3'].value}"))
        okh = all(str(ws.cell(row=5 + i, column=2).value) == h[0]
                  for i, h in enumerate(info["hawbs"]))
        chk.append(("交运单分单行齐全", bool(okh), ""))
    else:
        chk.append(("交运单存在", False, ""))
    return chk


def esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def safe_open(path):
    """打开文件/网址；服务器（无桌面）环境静默跳过。"""
    if os.environ.get('RUNPKG_NO_OPEN'):
        return
    try:
        os.startfile(path)  # Windows
    except AttributeError:
        import subprocess, webbrowser
        if str(path).startswith(('http://', 'https://')):
            webbrowser.open(path)
        elif os.name != 'nt':  # Linux 无桌面时 xdg-open 失败也不报错
            try:
                subprocess.Popen(['xdg-open', path],
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except Exception:
                pass
    except Exception:
        pass


def escq(s):
    """HTML 属性值转义（双引号也转），用于 input value 等。"""
    return esc(s).replace('"', '&quot;')


def build_report(eml_name, meta, masters, recs, model2cert, all_checks, gate_issues, ok):
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    color = "#1a7f37" if ok else "#b42318"
    banner = "全部校验通过，可以出单" if ok else "存在异常，禁止直接出单，按下方提示人工处理"
    h = [f"""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8">
<title>核对单 {esc(eml_name)}</title>
<style>
body{{font-family:'Microsoft YaHei',sans-serif;margin:24px;color:#222}}
h1{{font-size:20px}} .banner{{padding:10px 14px;color:#fff;background:{color};border-radius:6px;font-size:16px;margin:12px 0}}
table{{border-collapse:collapse;margin:8px 0 18px;font-size:13px}}
th,td{{border:1px solid #bbb;padding:4px 8px;text-align:left;vertical-align:top}}
th{{background:#f0f0f0}} .ok{{color:#1a7f37;font-weight:bold}} .bad{{color:#b42318;font-weight:bold}}
h2{{font-size:16px;margin:18px 0 6px}} .meta{{font-size:13px;color:#555}}
</style></head><body>
<h1>出货手续包核对单</h1>
<div class="meta">邮件：{esc(eml_name)}　|　生成时间：{now}　|　目的站 {esc(meta.get('dest','-'))}　航班 {esc(meta.get('flight','-'))}　运输日期 {esc(meta.get('date','-'))}</div>
<div class="banner">{banner}</div>"""]
    if gate_issues:
        h.append("<h2>异常与人工介入点</h2><table><tr><th>编号</th><th>问题</th></tr>")
        for i, iss in enumerate(gate_issues, 1):
            h.append(f"<tr><td>{i}</td><td class='bad'>{esc(iss)}</td></tr>")
        h.append("</table>")
    for master, info in masters.items():
        h.append(f"<h2>主单 {esc(master)}</h2>")
        h.append("<table><tr><th>HAWB</th><th>含电池</th><th>件数</th><th>毛重</th><th>机型</th></tr>")
        seen = set()
        hawb_ids = [x[0] for x in info["hawbs"]]
        for r in recs:
            if r["master"] == master and r["hawb"] in hawb_ids and r["hawb"] not in seen:
                seen.add(r["hawb"])
                h.append(f"<tr><td>{esc(r['hawb'])}</td><td>{'是' if r['is_battery'] else '否'}</td>"
                         f"<td>{r['pcs']}</td><td>{r['wt']}</td><td>{esc(','.join(r['models']) or '-')}</td></tr>")
        h.append("</table><table><tr><th>机型</th><th>鉴定证书编号</th></tr>")
        for m in info["models"]:
            c = model2cert.get(m)
            h.append(f"<tr><td>{m}</td><td>{esc(c['code']) if c else '<未匹配>'}</td></tr>")
        h.append("</table><table><tr><th>检查项</th><th>结果</th><th>说明</th></tr>")
        for name, good, detail in all_checks.get(master, []):
            h.append(f"<tr><td>{esc(name)}</td><td class='{'ok' if good else 'bad'}'>"
                     f"{'通过' if good else '不通过'}</td><td>{esc(detail)}</td></tr>")
        h.append("</table>")
    h.append("</body></html>")
    return "".join(h)


def analyze(emls):
    """解析全部邮件：正文/meta/分单记录/范围。G1 类问题（页面补不了）进 hard_issues。"""
    st = {'meta': {}, 'per_eml': [], 'certs': bw.load_certs(XLSX)}
    hard = []
    for eml in emls:
        body, msg = get_body(eml)
        if not body:
            hard.append(f"G1 邮件 {os.path.basename(eml)} 没有文本正文，无法解析。")
            st['per_eml'].append((eml, None, msg, [], None))
            continue
        meta, _ = extract_meta(body, msg)
        for k, v in meta.items():
            st['meta'].setdefault(k, v)
        recs = bw.parse_email(eml)
        sc = set(bw.top_masters(body)) if body else None
        if not recs:
            hard.append(f"G1 邮件 {os.path.basename(eml)} 解析不到「件数 计费 毛重」表格块，格式可能变了，需人工核对。")
        elif not bw.build_masters(recs, sc):
            hard.append(f"G1 邮件 {os.path.basename(eml)} 解析不到主单号（176-XXXXXXXX），需人工核对。")
        st['per_eml'].append((eml, body, msg, recs, sc))
    st['hard_issues'] = hard
    rebuild_state(st)
    return st


def rebuild_state(st):
    """由 certs + 各邮件解析结果重建 recs_all/model2cert/masters/entries（补充输入后重算）。
    cert_overrides（核对单人工改的证书映射）优先于 xlsx 自动匹配。"""
    st['cert_overrides'] = st.get('cert_overrides') or {}
    st['model2cert'] = {}
    for c in st['certs']:
        if bw.variant_cat(c['cn']) == 'phone':
            for mdl in c['models']:
                st['model2cert'][mdl] = c
    for c in st['certs']:
        for mdl in c['models']:
            st['model2cert'].setdefault(mdl, c)
    for mdl, c in st['cert_overrides'].items():
        st['model2cert'][mdl] = c
    st['recs_all'], st['masters_all'], st['entries_all'] = [], {}, {}
    for eml, body, msg, recs, sc in st['per_eml']:
        if not recs:
            continue
        st['recs_all'] += recs
        st['masters_all'].update(bw.build_masters(recs, sc))
        st['entries_all'].update(bw.build_entries(recs, sc, st['certs'], st['cert_overrides']))


def collect_gate_issues(st):
    """可复核闸门 G2-G5（G1 硬伤走 hard_issues）。"""
    issues = []
    meta = st['meta']
    if not meta.get('dest'):
        issues.append("G3 目的站：邮件里没有独立成行的三字母目的站（如 RUH），请在本页补充。")
    if not meta.get('flight'):
        issues.append("G3 航班：邮件里没有「EK309/Aug 16」样式的航班，请在本页补充。")
    if not meta.get('date'):
        issues.append("G3 运输日期：航班后的日期（如 2026-08-16）缺失，请在本页补充。")
    for eml, body, msg, recs, sc in st['per_eml']:
        for r in recs:
            if sc is not None and r['master'] not in sc:
                continue
            if r['master'] and r['hawb'] and (r['pcs'] is None or r['wt'] is None):
                issues.append(f"G2 分单 {r['hawb']} 的件数/毛重没解析出来，请在本页补充。")
            if r['master'] and r['is_battery'] and not r['models']:
                issues.append(f"G2 分单 {r['hawb']} 标了含电池但没提取到机型，请在本页补充。")
    for master, lst in st['entries_all'].items():
        for mdl, cat, c in lst:
            if not c:
                issues.append(
                    f"G5 主单 {master} 机型 {mdl} 没有{'数字式手机/5G' if cat == 'phone' else '手机模组'}品名变体行：请在本页补中英文品名，或补进鉴定汇总 xlsx 后重新校验。")
    for master, ent in st['masters_all'].items():
        miss_c = sorted(m for m in ent['models'] if m not in st['model2cert'])
        if miss_c:
            issues.append(
                f"G4 主单 {master} 机型 {miss_c} 在鉴定汇总 xlsx 无对应证书：在本页补证书编号/品名，或跑 chinasdg_query.py 写入 xlsx 后重新校验。")
        tdir = os.path.join(ROOT, '测试2', master + '模板')
        if not os.path.isdir(tdir) or not any('托书' in f for f in os.listdir(tdir)) \
                or not any(('运单副本' in f or '运单复本' in f) for f in os.listdir(tdir)):
            issues.append(
                f"G5 主单 {master} 缺模板：把该主单的 托书.pdf + 运单副本.pdf 放到 templates/测试2/{master}模板/ 后点按钮重新校验。")
    return issues


def apply_patch(st, patch):
    """合并页面补充输入：meta / 分单件重与机型 / 手工鉴定证书 / 删除分单与机型。"""
    m = patch.get('meta') or {}
    for k in ('dest', 'flight', 'date'):
        v = str(m.get(k) or '').strip()
        if v:
            st['meta'][k] = v
    # 托书顶部 CONSOLIDATION 声明（默认开）
    if 'consolidation' in m:
        st['meta']['consolidation'] = bool(m.get('consolidation'))
    elif 'consolidation' not in st['meta']:
        st['meta']['consolidation'] = True
    if st['meta'].get('flight'):
        st['meta']['carrier'] = st['meta']['flight'][:2]
    # 删除分单：按行号删除对应行（前端传 [{ri, hawb}]，重复 hawb 也只会删点击的那一行）
    del_recs = patch.get('del', {}).get('recs') or []
    if del_recs:
        rows = []  # 与 build_form_html 相同的渲染顺序：per_eml -> recs（跳过无主单/scope外）
        for eml, body, msg, recs, sc in st['per_eml']:
            for r in recs:
                if not (r['master'] and r['hawb']):
                    continue
                if sc is not None and r['master'] not in sc:
                    continue
                rows.append(r)
        del_idx = {int(d['ri']) for d in del_recs}
        for ri in sorted(del_idx, reverse=True):
            if 0 <= ri < len(rows):
                target = rows[ri]
                for eml, body, msg, recs, sc in st['per_eml']:
                    recs[:] = [r for r in recs if r is not target]
    # 删除机型：从所有分单 models 移除（证书行残留不影响，masters 不再引用）
    del_mdls = set(patch.get('del', {}).get('mdls') or [])
    if del_mdls:
        for eml, body, msg, recs, sc in st['per_eml']:
            for r in recs:
                r['models'] = [x for x in r['models'] if x not in del_mdls]
    for hawb, f in (patch.get('recs') or {}).items():
        for r in st['recs_all']:
            if r['hawb'] != hawb:
                continue
            v = str(f.get('master') or '').strip()
            if v:
                r['master'] = v
            v = str(f.get('hawb') or '').strip()
            if v:
                r['hawb'] = v
            v = str(f.get('is_battery') or '').strip()
            if v in ('1', '0'):
                r['is_battery'] = (v == '1')
            v = str(f.get('pcs') or '').strip()
            if v:
                try:
                    r['pcs'] = int(float(v))
                except ValueError:
                    pass
            v = str(f.get('wt') or '').strip()
            if v:
                try:
                    r['wt'] = float(v)
                except ValueError:
                    pass
            v = str(f.get('models') or '').strip()
            if v:
                r['models'] = list(dict.fromkeys(
                    x.strip() for x in re.split(r'[,，;；\s]+', v) if x.strip()))
    covered = {mdl for c in st['certs'] for mdl in c['models']}
    for mdl, f in (patch.get('certs') or {}).items():
        if mdl in covered:
            continue  # 鉴定汇总已有证书时以 xlsx 为准
        code = str(f.get('code') or '').strip()
        cn = str(f.get('cn') or '').strip()
        en = str(f.get('en') or '').strip()
        if not cn:
            continue
        models = list(dict.fromkeys(re.findall(r'[A-Z]\d{4}', cn) + [mdl]))
        st['certs'].append({'code': code, 'cn': cn, 'en': en,
                            'models': models, 'manual': True})


def apply_cert_edits(st, edits):
    """核对单上修改鉴定证书编号：edits = {机型: 新编号}。
    编号在鉴定汇总里能查到 → 改指到该证书行（品名随该行）；
    查不到 → 在当前证书上克隆新编号（品名不变）。返回改动说明列表。"""
    ov = st.setdefault('cert_overrides', {})
    notes = []
    for mdl, new_code in (edits or {}).items():
        new_code = str(new_code or '').strip()
        cur = st['model2cert'].get(mdl)
        if not new_code or (cur and new_code == str(cur['code']).strip()):
            continue
        hit = next((c for c in st['certs'] if str(c['code']).strip() == new_code), None)
        if hit is not None:
            ov[mdl] = hit
            notes.append(f"{mdl}：证书编号改为 {new_code}（品名改用该证书行）")
        elif cur is not None:
            clone = dict(cur)
            clone['code'] = new_code
            clone['manual'] = True
            ov[mdl] = clone
            notes.append(f"{mdl}：证书编号改为 {new_code}（汇总中无此编号，品名保持原证书行）")
    return notes


def do_generate(st):
    """闸门全过：设置 gp 参数，逐主单生成七份手续，返回生成后校验结果。"""
    meta = st['meta']
    gp.DEST = meta['dest']
    gp.FLIGHT = meta['flight']
    gp.CARRIER = meta.get('carrier') or meta['flight'][:2]
    gp.DATE_STR = meta['date']
    bw.inject_goods(gp, st['certs'])
    for c in st['certs']:  # 手工补的证书也进品名字典（回退路径用）
        if c.get('manual'):
            for mdl in c['models']:
                gp.CN_GOODS.setdefault(mdl, c['cn'])
                gp.EN_GOODS.setdefault(mdl, c['en'])
    gp.MASTERS = {}
    for m, ent in st['masters_all'].items():
        entries, codes, seen_cn = [], [], set()
        for _, _, c in st['entries_all'].get(m, []):
            if c and c['cn'] not in seen_cn:  # 手工证书可能同时命中两类变体，按中文品名去重
                seen_cn.add(c['cn'])
                entries.append((c['cn'], c['en']))
                if str(c['code']).strip() not in codes:  # 危险品确认单报告编号，与品名条目同序
                    codes.append(str(c['code']).strip())
        gp.MASTERS[m] = {'models': ent['models'],
                         'entries': entries,
                         'cert_codes': codes,
                         'parts': ent.get('parts', []),
                         'bpcs': ent.get('bpcs', 0),
                         'battery': bw.derive_battery_from([c for c, _ in entries]),
                         'pcs': sum(h[1] for h in ent['hawbs']),
                         'hawbs': ent['hawbs']}
    all_checks = {}
    consolidation = bool(meta.get('consolidation', True))  # 托书顶部 CONSOLIDATION 声明（默认开）
    for master, info in gp.MASTERS.items():
        print(f"---- 生成 {master} ----")
        gp.make_shengming(master, info)
        gp.make_anxian(master, info)
        gp.make_yingji(master, info)
        gp.fill_tuoshū(master, info, consolidation=consolidation)
        gp.fill_waybill(master, info)
        gp.fill_jiaoyun(master, info)
        gp.make_weixian(master, info)
        all_checks[master] = verify_master(master, info)
    return all_checks


def build_form_html(st, issues, port=0, ok_msg=None, base_url=None):
    """交互补充页：红框=缺失项，填完提交后合并重校验。
    port 给定时 JS 用绝对地址提交（页面存成本地文件直接打开也能用）；
    base_url 给定时提交地址用它拼 /submit（邮件守护令牌服务）；
    ok_msg 定制提交成功文案（邮件守护模式用「结果将回邮」）。"""
    meta = st['meta']
    if base_url:
        submit_url = base_url.rstrip('/') + '/submit'
        upload_url = base_url.rstrip('/') + '/upload'
    else:
        submit_url = f'http://127.0.0.1:{port}/submit' if port else '/submit'
        upload_url = f'http://127.0.0.1:{port}/upload' if port else '/upload'
    ok_msg = ok_msg or '闸门校验通过，即将打开「生成前核对单」——请核对信息（证书编号可直接修改），确认无误后点「确认生成」。'

    h = [f"""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8">
<title>信息补充 - 出货手续包</title>
<style>
body{{font-family:'Microsoft YaHei',sans-serif;margin:24px;color:#222}}
h1{{font-size:20px}} h2{{font-size:16px;margin:18px 0 6px}}
.banner{{padding:10px 14px;color:#fff;background:#b42318;border-radius:6px;font-size:15px;margin:12px 0}}
table{{border-collapse:collapse;margin:8px 0 18px;font-size:13px}}
th,td{{border:1px solid #bbb;padding:4px 8px;text-align:left;vertical-align:top}}
th{{background:#f0f0f0}}
input{{font-size:13px;padding:3px 5px;border:1px solid #999;border-radius:3px}}
input.miss{{background:#fdecea;border-color:#b42318}}
select{{font-size:13px;padding:3px 5px;border:1px solid #999;border-radius:3px}}
.delbtn{{font-size:12px;padding:3px 10px;background:#dc2626;color:#fff;border:none;border-radius:4px;cursor:pointer}}
.delbtn:hover{{background:#b91c1c}}
tr.del{{background:#fee2e2;color:#aaa}}
tr.del input,tr.del select{{background:#fee2e2}}
.tip{{font-size:12px;color:#666;margin:4px 0 10px;line-height:1.6}}
button{{font-size:15px;padding:8px 22px;background:#1a7f37;color:#fff;border:none;border-radius:6px;cursor:pointer;margin:10px 0}}
.bad{{color:#b42318;font-weight:bold}} .ok{{color:#1a7f37;font-weight:bold}}
#issues li{{margin:3px 0}}
</style></head><body>
<h1>出货手续包 — 信息补充页</h1>
<div class="tip">以下信息在订舱邮件/鉴定汇总里缺失，请补齐红色框项后点「补充并重新校验」；全部闸门通过后会自动生成手续包并打开核对单。</div>
<div id="issues"><div class="banner">当前缺 {len(issues)} 项信息/异常</div><ul>{''.join(f'<li>{esc(i)}</li>' for i in issues)}</ul></div>
<h2>① 基础信息（G3）</h2>
<table><tr>
<th>目的站（三字母）</th><td><input id="m-dest" value="{escq(meta.get('dest',''))}" size="10" class="{'miss' if not meta.get('dest') else ''}" placeholder="如 RUH"></td>
<th>航班</th><td><input id="m-flight" value="{escq(meta.get('flight',''))}" size="10" class="{'miss' if not meta.get('flight') else ''}" placeholder="如 EK309"></td>
<th>运输日期</th><td><input id="m-date" value="{escq(meta.get('date',''))}" size="14" class="{'miss' if not meta.get('date') else ''}" placeholder="如 2026-08-16"></td>
</tr><tr>
<th>托书声明</th><td colspan="3"><label style="font-weight:normal;font-size:13px;cursor:pointer"><input type="checkbox" id="m-consolidation" {'checked' if meta.get('consolidation', True) else ''}> 托书顶部加「CONSOLIDATION AS PER ATTACHED MANIFEST」（默认勾选）</label></td>
</tr></table>"""]
    h.append("<h2>② 分单件数/毛重/机型（G2）</h2>"
             "<div class='tip'>可修改主单/分单号、含电池标记、件数、毛重、机型；不需要的分单点「删除」移除。</div>"
             "<table><tr><th>主单</th><th>分单</th><th>含电池</th>"
             "<th>件数</th><th>毛重</th><th>机型（逗号分隔）</th><th>操作</th></tr>")
    ri = 0
    for eml, body, msg, recs, sc in st['per_eml']:
        for r in recs:
            if not (r['master'] and r['hawb']):
                continue
            if sc is not None and r['master'] not in sc:
                continue
            pcs_miss = 'miss' if r['pcs'] is None else ''
            wt_miss = 'miss' if r['wt'] is None else ''
            mdl_miss = 'miss' if r['is_battery'] and not r['models'] else ''
            bat_opts = ''.join(
                f"<option value='1'{' selected' if r['is_battery'] else ''}>是</option>"
                f"<option value='0'{' selected' if not r['is_battery'] else ''}>否</option>")
            h.append(f"<tr id='rec-{ri}'>"
                     f"<td><input data-ri='{ri}' data-rec='{escq(r['hawb'])}' data-field='master' size='14' value='{escq(r['master'])}'></td>"
                     f"<td><input data-ri='{ri}' data-rec='{escq(r['hawb'])}' data-field='hawb' size='14' value='{escq(r['hawb'])}'></td>"
                     f"<td><select data-ri='{ri}' data-rec='{escq(r['hawb'])}' data-field='is_battery'>{bat_opts}</select></td>"
                     f"<td><input data-ri='{ri}' data-rec='{escq(r['hawb'])}' data-field='pcs' size='4' class='{pcs_miss}' value='{'' if r['pcs'] is None else r['pcs']}'></td>"
                     f"<td><input data-ri='{ri}' data-rec='{escq(r['hawb'])}' data-field='wt' size='6' class='{wt_miss}' value='{'' if r['wt'] is None else r['wt']}'></td>"
                     f"<td><input data-ri='{ri}' data-rec='{escq(r['hawb'])}' data-field='models' size='28' class='{mdl_miss}' value='{escq(','.join(r['models']))}' placeholder='如 A2636,A2881'></td>"
                     f"<td><button type='button' class='delbtn' onclick='delRec({ri})'>删除</button></td></tr>")
            ri += 1
    h.append("</table>")
    allm = sorted({m for ent in st['masters_all'].values() for m in ent['models']})
    h.append("<h2>③ 机型鉴定信息（G4/G5）</h2>"
             "<div class='tip'>已匹配机型预填了鉴定汇总现值，一般不用改；标<b>需补充</b>的行必填。"
             "中文品名须以「数字式手机」/「5G数字移动电话机」开头（模组变体以「手机模组」开头），"
             "且包含「内置聚合物锂离子电池 + 电池型号、电压/容量/Wh」（托运人声明从这里反提电池型号）；"
             "英文品名以 DIGITAL CELLPHONE / 5G DIGITAL MOBILE PHONE FOR RADIOTELEPHONY 开头。</div>"
             "<table><tr><th>机型</th><th>证书编号</th><th>中文品名</th><th>英文品名</th><th>状态</th><th>操作</th></tr>")
    for mdl in allm:
        c = st['model2cert'].get(mdl)
        miss = '' if c else 'miss'
        h.append(f"<tr id='mdl-{mdl}'>"
                 f"<td>{mdl}</td>"
                 f"<td><input data-model='{mdl}' data-field='code' size='14' class='{miss}' value='{escq(c['code']) if c else ''}'></td>"
                 f"<td><input data-model='{mdl}' data-field='cn' size='48' class='{miss}' value='{escq(c['cn']) if c else ''}'></td>"
                 f"<td><input data-model='{mdl}' data-field='en' size='48' class='{miss}' value='{escq(c['en']) if c else ''}'></td>"
                 f"<td class={'ok' if c else 'bad'}>{'已匹配' if c else '需补充'}</td>"
                 f"<td><button type='button' class='delbtn' onclick='delMdl(\"{mdl}\")'>删除</button></td></tr>")
    h.append("</table>")
    h.append("<h2>④ 模板文件（G5，缺模板时可直接上传，也可发邮件附件）</h2>")
    miss_t = []
    for master in st['masters_all']:
        tdir = os.path.join(ROOT, '测试2', master + '模板')
        if not os.path.isdir(tdir) or not any('托书' in f for f in os.listdir(tdir)) \
                or not any(('运单副本' in f or '运单复本' in f) for f in os.listdir(tdir)):
            miss_t.append(master)
    if miss_t:
        h.append("<ul>" + "".join(
            f"<li class='bad'>主单 {m}：把 {m}托书.pdf + {m}运单副本.pdf 放到 templates/测试2/{m}模板/</li>"
            for m in miss_t) + "</ul>")
    else:
        h.append("<p class='ok'>模板齐全。</p>")
    h.append("""<p><input type="file" id="tpl" multiple accept="application/pdf,.pdf"> <button id="up" onclick="upload()">上传模板</button> <span id="uptip" style="font-size:12px;color:#666"></span></p>
    <p class="tip">文件名为「主单号+托书.pdf / 主单号+运单副本.pdf」（如 176-61333915托书.pdf），一次可传多个；入库后自动重新校验。</p>
    <button id="go" onclick="submit()">补充并重新校验</button>
<script>
var DEL={recs:{},mdls:{}};
function delRec(ri){
  var r=document.getElementById('rec-'+ri);
  if(!r)return;
  var hb=(r.querySelector('input[data-field=hawb]')||{}).value||r.getAttribute('data-hawb')||'';
  DEL.recs[ri]=hb;
  r.className='del';
}
function delMdl(m){DEL.mdls[m]=1;var r=document.getElementById('mdl-'+m);if(r)r.className='del';}
function upload(){
  var f=document.getElementById('tpl');
  if(!f.files.length){document.getElementById('uptip').textContent='请先选择 PDF 文件';return;}
  var fd=new FormData();
  for(var i=0;i<f.files.length;i++)fd.append('tpl',f.files[i]);
  var btn=document.getElementById('up');btn.disabled=true;btn.textContent='上传中…';
  fetch('__UPLOAD_URL__',{method:'POST',body:fd})
  .then(function(r){return r.json();})
  .then(function(j){
    btn.disabled=false;btn.textContent='上传模板';
    var tip=document.getElementById('uptip');
    if(j.saved&&j.saved.length){tip.textContent='已入库：'+j.saved.join('、');}
    else{tip.textContent='未识别到模板：文件名需含主单号+「托书」或「运单副本」';}
    if(j.remain&&j.remain.length){
      var box=document.getElementById('issues');
      box.innerHTML='<div class="banner">仍有 '+j.remain.length+' 项缺失/异常</div><ul>'
        +j.remain.map(function(i){return '<li>'+i.replace(/</g,'&lt;')+'</li>';}).join('')+'</ul>';
    }else if(j.confirm_url){
      window.location.href=j.confirm_url;
    }
  })
  .catch(function(e){
    btn.disabled=false;btn.textContent='上传模板';
    document.getElementById('uptip').textContent='上传失败：'+e;
  });
}
function collect(){
  var d={meta:{},recs:{},certs:{},del:{recs:[],mdls:[]}};
  d.meta.dest=document.getElementById('m-dest').value;
  d.meta.consolidation=document.getElementById('m-consolidation').checked;
  d.meta.flight=document.getElementById('m-flight').value;
  d.meta.date=document.getElementById('m-date').value;
  document.querySelectorAll('input[data-rec]').forEach(function(i){
    var ri=i.dataset.ri;
    if(DEL.recs[ri])return;
    var h=i.dataset.rec;
    if(!d.recs[h])d.recs[h]={};
    d.recs[h][i.dataset.field]=i.value;});
  document.querySelectorAll('select[data-rec]').forEach(function(i){
    var ri=i.dataset.ri;
    if(DEL.recs[ri])return;
    var h=i.dataset.rec;
    if(!d.recs[h])d.recs[h]={};
    d.recs[h][i.dataset.field]=i.value;});
  document.querySelectorAll('input[data-model]').forEach(function(i){
    if(DEL.mdls[i.dataset.model])return;
    if(!d.certs[i.dataset.model])d.certs[i.dataset.model]={};
    d.certs[i.dataset.model][i.dataset.field]=i.value;});
  d.del.recs=Object.keys(DEL.recs).map(function(ri){return {ri:parseInt(ri),hawb:DEL.recs[ri]};});
  d.del.mdls=Object.keys(DEL.mdls);
  return d;
}
function submit(){
  var btn=document.getElementById('go');
  btn.disabled=true;btn.textContent='校验中…';
  fetch('__SUBMIT_URL__',{method:'POST',body:JSON.stringify(collect())})
  .then(function(r){return r.json();})
  .then(function(j){
    var box=document.getElementById('issues');
    if(j.ok){
      box.innerHTML='<div class="banner" style="background:#1a7f37">__OK_MSG__</div>';
      document.querySelectorAll('input').forEach(function(e){e.disabled=true;});
      btn.style.display='none';
    }else if(j.confirm&&j.confirm_url){
      window.location.href=j.confirm_url;
    }else{
      box.innerHTML='<div class="banner">仍有 '+j.issues.length+' 项缺失/异常，请继续补充</div><ul>'
        +j.issues.map(function(i){return '<li>'+i.replace(/</g,'&lt;')+'</li>';}).join('')+'</ul>';
      btn.disabled=false;btn.textContent='补充并重新校验';
      window.scrollTo(0,0);
    }
  })
  .catch(function(e){
    document.getElementById('issues').innerHTML='<div class="banner">提交失败：'+e+'</div>';
    btn.disabled=false;btn.textContent='补充并重新校验';
  });
}
</script></body></html>""")
    return "".join(h).replace('__SUBMIT_URL__', submit_url).replace('__UPLOAD_URL__', upload_url).replace('__OK_MSG__', ok_msg)


def serve_fill_form(st, issues):
    """闸门不过且可页面补充时：起 127.0.0.1 临时服务开补充页；返回是否生成成功。
    多线程服务 + 连接超时：防代理/安全软件只连不发请求把单线程服务卡死；
    页面同时落盘成本地文件，浏览器被代理拦住打不开 URL 时直接打开文件提交。"""
    import http.server
    import json as _json
    state = {'done': False, 'ok': False, 'confirm': False}
    page = None  # 服务起好拿到端口后再渲染（JS 提交地址含端口）

    class Handler(http.server.BaseHTTPRequestHandler):
        protocol_version = 'HTTP/1.1'
        timeout = 60  # 挂住不发的连接（代理探测等）60s 自动断开，不阻塞其他请求

        def log_message(self, fmt, *args):
            pass

        def _send(self, code, body, ctype):
            data = body.encode('utf-8')
            self.send_response(code)
            self.send_header('Content-Type', ctype)
            self.send_header('Content-Length', str(len(data)))
            self.send_header('Access-Control-Allow-Origin', '*')  # file:// 打开备份页也能提交
            self.end_headers()
            self.wfile.write(data)

        def do_GET(self):
            if self.path in ('/', '/index.html'):
                self._send(200, page, 'text/html; charset=utf-8')
            else:
                self._send(404, 'not found', 'text/plain')

        def do_POST(self):
            if self.path != '/submit':
                return self._send(404, 'not found', 'text/plain')
            n = int(self.headers.get('Content-Length') or 0)
            try:
                patch = _json.loads(self.rfile.read(n).decode('utf-8'))
            except Exception:
                return self._send(400, _json.dumps({'ok': False, 'issues': ['补充输入解析失败']},
                                                   ensure_ascii=False), 'application/json')
            try:
                st['certs'] = bw.load_certs(XLSX)  # 用户可能补过 xlsx，重读
            except Exception:
                pass
            apply_patch(st, patch)
            rebuild_state(st)
            remain = collect_gate_issues(st)
            if remain:
                print("\n".join(remain))
                return self._send(200, _json.dumps({'ok': False, 'issues': remain},
                                                   ensure_ascii=False), 'application/json')
            # 闸门全过：不直接生成，转入确认页等人工确认后才生成
            print("\n[闸门] 全部通过，转入确认页，等用户确认后再生成。")
            state['done'], state['confirm'] = True, True
            self._send(200, _json.dumps({'ok': True, 'confirm': True},
                                        ensure_ascii=False), 'application/json')

    srv = http.server.ThreadingHTTPServer(('127.0.0.1', 0), Handler)
    srv.daemon_threads = True
    port = srv.server_address[1]
    page = build_form_html(st, issues, port)
    # 备份页落盘：浏览器打不开 URL（如代理拦截）时直接打开该文件提交，效果相同
    os.makedirs(OUT, exist_ok=True)
    fp = os.path.join(OUT, '补充页_%s.html' % datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))
    open(fp, 'w', encoding='utf-8').write(page)
    print(f"[补充] 信息有缺失，请在页面补齐后提交（本窗口保持开启）")
    print(f"  页面地址：http://127.0.0.1:{port}/")
    print(f"  若浏览器一直加载打不开，直接打开本地文件：{fp}")
    print("\n".join(issues))
    safe_open(f'http://127.0.0.1:{port}/')
    try:
        srv.timeout = 0.5  # handle_request 超时即返回，多线程下才能及时检查 done 标志退出
        while not state['done']:
            srv.handle_request()
    except KeyboardInterrupt:
        print("\n[中断] 已取消，未生成手续包。")
    srv.server_close()
    return 'confirm' if state.get('confirm') else None


def build_confirm_html(st, port=0, base_url=None):
    """生成前核对单：展示全部解析结果，机型鉴定证书编号可改，确认后才生成。"""
    meta = st['meta']
    if base_url:
        submit_url = base_url.rstrip('/') + '/confirm_submit'
        cancel_url = base_url.rstrip('/') + '/cancel'
    else:
        submit_url = f'http://127.0.0.1:{port}/confirm_submit' if port else '/confirm_submit'
        cancel_url = f'http://127.0.0.1:{port}/cancel' if port else '/cancel'
    h = [f"""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8">
<title>生成前核对确认 - 出货手续包</title>
<style>
body{{font-family:'Microsoft YaHei',sans-serif;margin:24px;color:#222}}
h1{{font-size:20px}} h2{{font-size:16px;margin:18px 0 6px}}
.banner{{padding:10px 14px;color:#fff;background:#1a7f37;border-radius:6px;font-size:15px;margin:12px 0}}
table{{border-collapse:collapse;margin:8px 0 18px;font-size:13px}}
th,td{{border:1px solid #bbb;padding:4px 8px;text-align:left;vertical-align:top}}
th{{background:#f0f0f0}}
input{{font-size:13px;padding:3px 5px;border:1px solid #999;border-radius:3px}}
.tip{{font-size:12px;color:#666;margin:4px 0 10px;line-height:1.6}}
.goods{{font-size:12px;color:#333;max-width:430px}}
button{{font-size:15px;padding:8px 22px;color:#fff;border:none;border-radius:6px;cursor:pointer;margin:10px 8px 10px 0}}
#go{{background:#1a7f37}} #cancel{{background:#6b7280}}
.bad{{color:#b42318;font-weight:bold}}
</style></head><body>
<h1>出货手续包 — 生成前核对单</h1>
<div class="tip">闸门校验已全部通过。<b>请逐项核对后再点「确认生成」，未确认不会生成任何文件。</b>
鉴定证书编号可在输入框里直接修改：改后的编号若在鉴定汇总里能查到，该机型的品名会同步改用对应证书行；查不到则只改编号、品名不变。</div>
<div id="box"><div class="banner">闸门全部通过，待人工确认　|　目的站 {esc(meta.get('dest','-'))}　航班 {esc(meta.get('flight','-'))}　运输日期 {esc(meta.get('date','-'))}</div></div>"""]
    for master, ent in st['masters_all'].items():
        h.append(f"<h2>主单 {esc(master)}（总件数 {ent.get('pcs', '-')}，含电池分单 {ent.get('bpcs', 0)} 件）</h2>")
        h.append("<table><tr><th>HAWB</th><th>含电池</th><th>件数</th><th>毛重</th><th>机型</th></tr>")
        for r in st['recs_all']:
            if r['master'] != master or not r['hawb']:
                continue
            h.append(f"<tr><td>{esc(r['hawb'])}</td><td>{'是' if r['is_battery'] else '否'}</td>"
                     f"<td>{r['pcs'] if r['pcs'] is not None else '-'}</td>"
                     f"<td>{r['wt'] if r['wt'] is not None else '-'}</td>"
                     f"<td>{esc(','.join(r['models']) or '-')}</td></tr>")
        h.append("</table><table><tr><th>机型</th><th>鉴定证书编号（可修改）</th>"
                 "<th>中文品名（按当前证书行）</th><th>英文品名（按当前证书行）</th></tr>")
        for mdl in ent['models']:
            c = st['model2cert'].get(mdl)
            h.append(f"<tr><td>{mdl}</td>"
                     f"<td><input data-cert='{mdl}' size='20' value='{escq(c['code']) if c else ''}'></td>"
                     f"<td class='goods'>{esc(c['cn']) if c else '<未匹配>'}</td>"
                     f"<td class='goods'>{esc(c['en']) if c else '<未匹配>'}</td></tr>")
        h.append("</table>")
    h.append(f"""<button id="go" onclick="go()">确认生成</button>
<button id="cancel" onclick="cancel()">取消（不生成）</button>
<script>
function collect(){{
  var d={{certs:{{}}}};
  document.querySelectorAll('input[data-cert]').forEach(function(i){{d.certs[i.dataset.cert]=i.value;}});
  return d;
}}
function go(){{
  var btn=document.getElementById('go');
  btn.disabled=true;btn.textContent='正在生成…';
  fetch('{submit_url}',{{method:'POST',body:JSON.stringify(collect())}})
  .then(function(r){{return r.json();}})
  .then(function(j){{
    var box=document.getElementById('box');
    if(j.ok){{
      box.innerHTML='<div class="banner">已生成完成，正式核对单（含逐项校验结果）已打开，可关闭本页。</div>'
        +((j.notes&&j.notes.length)?('<div class="banner" style="background:#b45309">证书改动：'+j.notes.join('；')+'</div>'):'')
        +(j.bad>0?('<div class="banner" style="background:#b42318">生成后校验有 '+j.bad+' 项不通过，请按正式核对单人工复核。</div>'):'');
      document.querySelectorAll('input').forEach(function(e){{e.disabled=true;}});
      btn.style.display='none';document.getElementById('cancel').style.display='none';
    }}else{{
      box.innerHTML='<div class="banner" style="background:#b42318">无法生成：'+j.issues.length+' 项问题</div><ul>'
        +j.issues.map(function(i){{return '<li class="bad">'+i.replace(/</g,'&lt;')+'</li>';}}).join('')+'</ul>';
      btn.disabled=false;btn.textContent='确认生成';
    }}
  }})
  .catch(function(e){{
    document.getElementById('box').innerHTML='<div class="banner" style="background:#b42318">提交失败：'+e+'（服务窗口可能已关闭）</div>';
    btn.disabled=false;btn.textContent='确认生成';
  }});
}}
function cancel(){{
  fetch('{cancel_url}',{{method:'POST'}}).catch(function(){{}});
  document.getElementById('box').innerHTML='<div class="banner" style="background:#6b7280">已取消，未生成手续包，可关闭本页。</div>';
  document.getElementById('go').style.display='none';document.getElementById('cancel').style.display='none';
}}
</script></body></html>""")
    return "".join(h)


def serve_confirm(st):
    """闸门全过后开「生成前核对单」：人工核对、可改证书编号，确认后才生成；取消不生成。返回是否生成成功。"""
    import http.server
    import json as _json
    state = {'done': False, 'ok': False}
    page = None

    class Handler(http.server.BaseHTTPRequestHandler):
        protocol_version = 'HTTP/1.1'
        timeout = 60

        def log_message(self, fmt, *args):
            pass

        def _send(self, code, body, ctype):
            data = body.encode('utf-8')
            self.send_response(code)
            self.send_header('Content-Type', ctype)
            self.send_header('Content-Length', str(len(data)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(data)

        def do_GET(self):
            if self.path in ('/', '/index.html'):
                self._send(200, page, 'text/html; charset=utf-8')
            else:
                self._send(404, 'not found', 'text/plain')

        def do_POST(self):
            if self.path == '/cancel':
                print("\n[取消] 用户在确认页点了取消，未生成手续包。")
                state['done'] = True
                return self._send(200, _json.dumps({'cancelled': True},
                                                   ensure_ascii=False), 'application/json')
            if self.path != '/confirm_submit':
                return self._send(404, 'not found', 'text/plain')
            n = int(self.headers.get('Content-Length') or 0)
            try:
                patch = _json.loads(self.rfile.read(n).decode('utf-8'))
            except Exception:
                return self._send(400, _json.dumps({'ok': False, 'issues': ['确认输入解析失败']},
                                                   ensure_ascii=False), 'application/json')
            notes = apply_cert_edits(st, patch.get('certs') or {})
            for x in notes:
                print(f"  [证书改动] {x}")
            rebuild_state(st)
            remain = collect_gate_issues(st)
            if remain:
                print("\n".join(remain))
                return self._send(200, _json.dumps({'ok': False, 'issues': remain},
                                                   ensure_ascii=False), 'application/json')
            try:
                all_checks = do_generate(st)
            except Exception as e:
                return self._send(200, _json.dumps({'ok': False, 'issues': [f'生成报错：{e}']},
                                                   ensure_ascii=False), 'application/json')
            bad = [(m, nm, d) for m, rows in all_checks.items()
                   for nm, good, d in rows if not good]
            ok = not bad
            eml_name = os.path.basename(st['per_eml'][0][0])
            report = build_report(eml_name, st['meta'], gp.MASTERS, st['recs_all'],
                                  st['model2cert'], all_checks, [], ok)
            rp = os.path.join(OUT, '核对单_%s.html' % datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))
            open(rp, 'w', encoding='utf-8').write(report)
            for m, nm, d in bad:
                print(f"  [不通过] {m} {nm} {d}")
            print(f"\n[完成] {len(gp.MASTERS)} 个主单 × 7 份手续已生成。正式核对单：{rp}")
            safe_open(rp)
            state['done'], state['ok'] = True, ok
            self._send(200, _json.dumps({'ok': True, 'bad': len(bad), 'notes': notes},
                                        ensure_ascii=False), 'application/json')

    srv = http.server.ThreadingHTTPServer(('127.0.0.1', 0), Handler)
    srv.daemon_threads = True
    port = srv.server_address[1]
    page = build_confirm_html(st, port)
    os.makedirs(OUT, exist_ok=True)
    fp = os.path.join(OUT, '核对单_待确认_%s.html' % datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))
    open(fp, 'w', encoding='utf-8').write(page)
    print("[确认] 闸门全部通过。请核对信息（证书编号可直接修改），点「确认生成」后才会生成手续包。")
    print(f"  确认页地址：http://127.0.0.1:{port}/")
    print(f"  若页面打不开，直接打开本地文件：{fp}")
    print("  点「取消」或在命令行按 Ctrl+C 可放弃生成（本窗口需保持开启）")
    safe_open(f'http://127.0.0.1:{port}/')
    try:
        srv.timeout = 0.5
        while not state['done']:
            srv.handle_request()
    except KeyboardInterrupt:
        print("\n[中断] 已取消，未生成手续包。")
    srv.server_close()
    return state['ok']


def finish(st, eml_name, tag=''):
    """生成七份手续 + 正式核对单落盘，返回 (all_checks, bad, ok, report_path)。"""
    all_checks = do_generate(st)
    bad = [(m, n, d) for m, rows in all_checks.items() for n, good, d in rows if not good]
    ok = not bad
    report = build_report(eml_name, st['meta'], gp.MASTERS, st['recs_all'],
                          st['model2cert'], all_checks, [], ok)
    os.makedirs(OUT, exist_ok=True)
    rp = os.path.join(OUT, '核对单_%s%s.html' % (datetime.datetime.now().strftime('%Y%m%d_%H%M%S'), tag))
    open(rp, 'w', encoding='utf-8').write(report)
    for m, n, d in bad:
        print(f"  [不通过] {m} {n} {d}")
    if ok:
        print(f"\n[完成] {len(gp.MASTERS)} 个主单 × 7 份手续，全部校验通过。报告：{rp}")
    else:
        print(f"\n[异常] 生成后校验有 {len(bad)} 项不通过，请人工复核。报告：{rp}")
    return all_checks, bad, ok, rp


def run_headless(emls, generate=True):
    """服务器/邮件守护模式：不开页面、不打开文件、不等人工确认。
    generate=False 只探测闸门不生成（邮件守护先建补充页，补齐后 finish 一步生成）。
    返回 dict：
      status='hard'      G1 硬伤（邮件解析不了），issues 为问题列表
      status='need_fill' 闸门缺信息，issues + st 交给调用方（邮件守护开令牌补充页）
      status='done'      已生成，含 all_checks/ok/report_path/masters
    """
    st = analyze(emls)
    if st['hard_issues']:
        return {'status': 'hard', 'issues': st['hard_issues'], 'st': st}
    issues = collect_gate_issues(st)
    if issues:
        return {'status': 'need_fill', 'issues': issues, 'st': st}
    if not generate:
        return {'status': 'ready', 'st': st}
    eml_name = os.path.basename(emls[0])
    try:
        all_checks, bad, ok, rp = finish(st, eml_name)
    except Exception as e:
        return {'status': 'hard', 'issues': [f'生成报错：{e}'], 'st': st}
    return {'status': 'done', 'all_checks': all_checks, 'ok': ok,
            'report_path': rp, 'masters': list(gp.MASTERS.keys()), 'st': st}


def main():
    args = [a for a in sys.argv[1:] if a.strip()]
    if not args:
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            p = filedialog.askopenfilename(
                title="选择订舱邮件",
                filetypes=[("订舱邮件", "*.eml"), ("所有文件", "*.*")])
            root.destroy()
            args = [p] if p else []
        except Exception:
            pass
    if not args:
        print("用法：把 .eml 订舱邮件拖到 一键出货手续包.bat 上运行。")
        return 2

    st = analyze(args)
    if st['hard_issues']:
        # G1 类问题页面补不了：仍出静态异常报告
        report = build_report(os.path.basename(args[0]), st['meta'], st['masters_all'],
                              st['recs_all'], st['model2cert'], {}, st['hard_issues'], False)
        os.makedirs(OUT, exist_ok=True)
        rp = os.path.join(OUT, '异常报告_%s.html' % datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))
        open(rp, 'w', encoding='utf-8').write(report)
        print("\n".join(st['hard_issues']))
        print(f"\n[异常] 共 {len(st['hard_issues'])} 项，未生成手续包。报告：{rp}")
        safe_open(rp)
        return 2

    # 闸门不过（可补充类）：开补充页交互补齐，补齐后转确认页
    issues = collect_gate_issues(st)
    if issues and serve_fill_form(st, issues) != 'confirm':
        return 2

    # 闸门全过：先出「生成前核对单」等人工确认（可改证书编号），确认后才生成
    if not os.environ.get('RUNPKG_AUTO_CONFIRM'):
        ok = serve_confirm(st)
        if not ok:
            print("\n[未生成] 已取消或未确认，未生成手续包。")
        return 0 if ok else 2

    # RUNPKG_AUTO_CONFIRM：跳过确认页直接生成（自动化/测试用）
    print("[确认] 检测到 RUNPKG_AUTO_CONFIRM，跳过人工确认直接生成。")
    _, bad, ok, rp = finish(st, os.path.basename(args[0]))
    safe_open(rp)
    return 0 if ok else 2


if __name__ == "__main__":
    sys.exit(main())
